import urllib.request
import urllib.parse
import json
import time
import urllib.error
import csv
import os
import re
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Border, Side, PatternFill
import matplotlib.pyplot as plt
import matplotlib.dates as mdates

# --- Configuration ---
DISCORD_WEBHOOK_URL = "" 
ALERT_THRESHOLD_PERCENT = 10.0 

def clean_txt(file_path):
    if not os.path.exists(file_path):
        # We don't print error here, we handle it in main
        return

    with open(file_path, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    seen = set()
    unique_lines = []
    for line in lines:
        stripped = line.strip()
        if stripped and stripped not in seen:
            seen.add(stripped)
            unique_lines.append(line)
        elif not stripped:
             pass

    # Write back
    with open(file_path, 'w', encoding='utf-8') as f:
        f.writelines(unique_lines)

# --- Magic: The Gathering (Scryfall) ---
def get_mtg_data(card_line):
    # Try to parse name and collector number
    match = re.search(r'^(.*?)\s+#\s*(\S+)$', card_line)
    
    queries = []
    clean_name = card_line
    
    if match:
        name_part = match.group(1).strip()
        number_part = match.group(2).strip()
        clean_name = re.sub(r'\s*\(.*?\)', '', name_part).strip()
        number_stripped = number_part.lstrip('0')
        
        queries.append(f'name:"{clean_name}" cn:"{number_part}"')
        queries.append(f'name:"{clean_name}" cn:"{number_stripped}"')
    elif ' - ' in card_line:
        # Try to handle "Name - Set" format
        parts = card_line.split(' - ')
        clean_name = parts[0].strip()
        set_info = parts[1].strip()
        queries.append(f'name:"{clean_name}" s:"{set_info}"')
        queries.append(f'name:"{clean_name}"') # Fallback to just name
    else:
        clean_name = card_line
    
    # Add fuzzy search as last resort
    encoded_name = urllib.parse.quote(clean_name)
    fuzzy_url = f"https://api.scryfall.com/cards/named?fuzzy={encoded_name}"
    
    # Try specific queries first
    for q in queries:
        encoded_query = urllib.parse.quote(q)
        url = f"https://api.scryfall.com/cards/search?q={encoded_query}"
        try:
            with urllib.request.urlopen(url) as response:
                data = json.loads(response.read().decode())
                if data.get('total_cards', 0) > 0:
                    card_data = data['data'][0]
                    return parse_scryfall_data(card_data, clean_name)
        except Exception:
            continue

    # Fallback to fuzzy search
    try:
        with urllib.request.urlopen(fuzzy_url) as response:
            data = json.loads(response.read().decode())
            return parse_scryfall_data(data, clean_name)
    except Exception as e:
        print(f"[MTG] Error fetching {card_line}: {e}")
        return None

def parse_scryfall_data(data, default_name):
    prices = data.get('prices', {})
    price = prices.get('usd')
    if not price:
        price = prices.get('usd_foil')
    
    name = data.get('name', default_name)
    if 'flavor_name' in data:
        name = data['flavor_name']

    image_url = ""
    if 'image_uris' in data:
        image_url = data['image_uris'].get('normal', '')
    elif 'card_faces' in data and 'image_uris' in data['card_faces'][0]:
         image_url = data['card_faces'][0]['image_uris'].get('normal', '')

    return {
        'game': 'MTG',
        'name': name,
        'set': data.get('set_name', 'Unknown'),
        'price': price if price else 'N/A',
        'image': image_url,
        'uri': data.get('scryfall_uri', '#')
    }

# --- Yu-Gi-Oh! (YGOPRODeck) ---
def get_yugioh_data(card_line):
    # YGO API is simpler: exact name match usually
    # Strip (...) for search, as API expects exact card name
    clean_name = re.sub(r'\s*\(.*?\)', '', card_line).strip()
    
    encoded_name = urllib.parse.quote(clean_name)
    url = f"https://db.ygoprodeck.com/api/v7/cardinfo.php?name={encoded_name}"
    
    try:
        with urllib.request.urlopen(url) as response:
            data = json.loads(response.read().decode())
            if data.get('data'):
                card_data = data['data'][0]
                return parse_ygo_data(card_data, clean_name)
    except urllib.error.HTTPError as e:
        # Try fuzzy/partial match if exact fails
        if e.code == 400: 
             return get_yugioh_data_fuzzy(clean_name)
        print(f"[YGO] Error fetching {card_line}: HTTP {e.code}")
        return None
    except Exception as e:
        print(f"[YGO] Error fetching {card_line}: {e}")
        return None
    return None

def get_yugioh_data_fuzzy(clean_name):
    encoded_name = urllib.parse.quote(clean_name)
    url = f"https://db.ygoprodeck.com/api/v7/cardinfo.php?fname={encoded_name}"
    try:
        with urllib.request.urlopen(url) as response:
            data = json.loads(response.read().decode())
            if data.get('data'):
                # Return first match
                card_data = data['data'][0]
                return parse_ygo_data(card_data, clean_name)
    except Exception:
        return None
    return None

def parse_ygo_data(card_data, default_name):
    # Price (TCGPlayer)
    price = 'N/A'
    if 'card_prices' in card_data and len(card_data['card_prices']) > 0:
        price = card_data['card_prices'][0].get('tcgplayer_price', 'N/A')
    
    # Image
    image_url = ""
    if 'card_images' in card_data and len(card_data['card_images']) > 0:
        image_url = card_data['card_images'][0].get('image_url', '')

    # Set (First set code if available, or just "Yu-Gi-Oh!")
    set_name = "Yu-Gi-Oh!"
    if 'card_sets' in card_data and len(card_data['card_sets']) > 0:
        set_name = card_data['card_sets'][0].get('set_name', 'Unknown Set')

    return {
        'game': 'YGO',
        'name': card_data.get('name', default_name),
        'set': set_name,
        'price': str(price),
        'image': image_url,
        'uri': f"https://db.ygoprodeck.com/card/?search={urllib.parse.quote(card_data.get('name'))}" 
    }


def update_history_and_graph(total_value):
    history_file = 'price_history.csv'
    today = datetime.now().strftime('%Y-%m-%d')
    
    # Read existing
    history = []
    previous_value = 0.0
    
    if os.path.exists(history_file):
        with open(history_file, 'r', newline='') as f:
            reader = csv.reader(f)
            history = list(reader)
            for row in reversed(history):
                if row[0] != today:
                    try:
                        previous_value = float(row[1])
                        break
                    except: pass
    
    updated = False
    for row in history:
        if row and row[0] == today:
            row[1] = f"{total_value:.2f}"
            updated = True
            break
    
    if not updated:
        history.append([today, f"{total_value:.2f}"])
        
    history.sort(key=lambda x: x[0])
    
    with open(history_file, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(history)
        
    if len(history) > 0:
        dates = [datetime.strptime(row[0], '%Y-%m-%d') for row in history]
        values = [float(row[1]) for row in history]
        
        plt.figure(figsize=(10, 5))
        plt.plot(dates, values, marker='o', linestyle='-', color='#6c5ce7', linewidth=2)
        plt.title('Total Collection Value', fontsize=14)
        plt.xlabel('Date')
        plt.ylabel('Value (USD)')
        plt.grid(True, linestyle='--', alpha=0.7)
        plt.gca().xaxis.set_major_formatter(mdates.DateFormatter('%m/%d'))
        plt.gcf().autofmt_xdate()
        plt.savefig('history_graph.png', bbox_inches='tight')
        plt.close()

    if previous_value > 0 and DISCORD_WEBHOOK_URL:
        percent_change = ((total_value - previous_value) / previous_value) * 100
        if percent_change >= ALERT_THRESHOLD_PERCENT:
            send_discord_alert(total_value, percent_change)

def send_discord_alert(current_value, percent_change):
    try:
        data = {
            "content": f"🚀 **Collection Value Alert!**\nYour collection is now worth **${current_value:.2f}**.\nThat's a **{percent_change:.1f}%** increase since the last check!"
        }
        req = urllib.request.Request(
            DISCORD_WEBHOOK_URL, 
            data=json.dumps(data).encode('utf-8'),
            headers={'Content-Type': 'application/json'}
        )
        urllib.request.urlopen(req)
        print("Discord alert sent!")
    except Exception as e:
        print(f"Failed to send Discord alert: {e}")

def generate_html_report(collected_data, total_value, total_profit_loss):
    today = datetime.now().strftime('%Y-%m-%d')
    pl_color = "#27ae60" if total_profit_loss >= 0 else "#c0392b"
    pl_sign = "+" if total_profit_loss >= 0 else ""
    
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>Multi-TCG Tracker</title>
        <style>
            body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background-color: #f4f4f9; color: #333; margin: 0; padding: 20px; }}
            .container {{ max-width: 1200px; margin: 0 auto; }}
            .header {{ background: white; padding: 20px; border-radius: 10px; box-shadow: 0 2px 5px rgba(0,0,0,0.1); display: flex; justify-content: space-between; align-items: center; margin-bottom: 20px; }}
            .header h1 {{ margin: 0; font-size: 24px; }}
            .stats {{ text-align: right; }}
            .total-value {{ font-size: 32px; font-weight: bold; color: #2c3e50; }}
            .profit-loss {{ font-size: 18px; font-weight: bold; color: {pl_color}; }}
            .graph-container {{ background: white; padding: 20px; border-radius: 10px; box-shadow: 0 2px 5px rgba(0,0,0,0.1); margin-bottom: 20px; text-align: center; }}
            .graph-container img {{ max-width: 100%; height: auto; }}
            .grid {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(200px, 1fr)); gap: 20px; }}
            .card {{ background: white; border-radius: 10px; overflow: hidden; box-shadow: 0 2px 5px rgba(0,0,0,0.1); transition: transform 0.2s; position: relative; }}
            .card:hover {{ transform: translateY(-5px); box-shadow: 0 5px 15px rgba(0,0,0,0.2); }}
            .card-img {{ width: 100%; height: 300px; object-fit: contain; background: #eee; }}
            .card-info {{ padding: 15px; }}
            .card-name {{ font-weight: bold; margin-bottom: 5px; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }}
            .card-set {{ color: #7f8c8d; font-size: 12px; margin-bottom: 10px; }}
            .card-price {{ font-size: 18px; color: #2980b9; font-weight: bold; }}
            .card-pl {{ font-size: 14px; font-weight: bold; margin-top: 5px; }}
            .pl-green {{ color: #27ae60; }}
            .pl-red {{ color: #c0392b; }}
            .card-qty {{ font-size: 12px; color: #95a5a6; float: right; }}
            .badge {{ position: absolute; top: 10px; right: 10px; padding: 5px 10px; border-radius: 15px; color: white; font-weight: bold; font-size: 12px; box-shadow: 0 2px 4px rgba(0,0,0,0.2); }}
            .badge-MTG {{ background-color: #f39c12; }}
            .badge-YGO {{ background-color: #8e44ad; }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <div>
                    <h1>My Collection</h1>
                    <p>Last Updated: {today}</p>
                </div>
                <div class="stats">
                    <div class="total-value">${total_value:.2f}</div>
                    <div class="profit-loss">P/L: {pl_sign}${total_profit_loss:.2f}</div>
                </div>
            </div>
            
            <div class="graph-container">
                <h2>Value History</h2>
                <img src="history_graph.png" alt="Value History Graph">
            </div>
            
            <div class="grid">
    """
    
    for item in collected_data:
        data = item['data']
        price_str = item['price_str']
        qty = item['quantity']
        profit_loss = item['profit_loss']
        game = data['game']
        
        pl_html = ""
        if profit_loss is not None:
            color_class = "pl-green" if profit_loss >= 0 else "pl-red"
            sign = "+" if profit_loss >= 0 else ""
            pl_html = f'<div class="card-pl {color_class}">{sign}${profit_loss:.2f}</div>'

        img_src = data['image'] if data['image'] else "https://via.placeholder.com/250x350?text=No+Image"
        
        html += f"""
                <div class="card" onclick="window.open('{data['uri']}', '_blank')" style="cursor: pointer;">
                    <span class="badge badge-{game}">{game}</span>
                    <img src="{img_src}" class="card-img" alt="{data['name']}">
                    <div class="card-info">
                        <div class="card-name" title="{data['name']}">{data['name']}</div>
                        <div class="card-set">{data['set']}</div>
                        <div class="card-price">
                            ${price_str}
                            <span class="card-qty">x{qty}</span>
                        </div>
                        {pl_html}
                    </div>
                </div>
        """
        
    html += """
            </div>
        </div>
    </body>
    </html>
    """
    
    with open('index.html', 'w', encoding='utf-8') as f:
        f.write(html)
    print("Report generated: index.html")

def process_file(filename, fetch_func, game_name):
    clean_txt(filename)
    try:
        with open(filename, 'r') as f:
            cards = [line.strip() for line in f if line.strip()]
    except FileNotFoundError:
        print(f"Note: {filename} not found. Skipping {game_name}.")
        return []

    results = []
    print(f"Fetching {game_name} prices for {len(cards)} cards...")
    
    for card in cards:
        quantity = 1
        bought_price = None
        
        if '|' in card:
            parts = card.split('|')
            card = parts[0].strip()
            try:
                bought_price = float(parts[1].strip())
            except ValueError:
                pass
        
        clean_card_line = card
        qty_match = re.search(r'^(\d+)x\s+(.*)', card)
        if qty_match:
            quantity = int(qty_match.group(1))
            clean_card_line = qty_match.group(2).strip()

        data = fetch_func(clean_card_line)
        if data:
            unit_price_str = data['price']
            total_price_str = unit_price_str
            profit_loss = None
            
            sort_val = -1.0
            if unit_price_str != 'N/A':
                try:
                    unit_price = float(unit_price_str)
                    total_price = unit_price * quantity
                    total_price_str = f"{total_price:.2f}"
                    sort_val = float(total_price_str)
                    
                    if bought_price is not None:
                        total_bought = bought_price * quantity
                        profit_loss = total_price - total_bought
                except ValueError:
                    pass

            results.append({
                'data': data,
                'quantity': quantity,
                'price_str': total_price_str,
                'profit_loss': profit_loss,
                'sort_val': sort_val
            })
            print(f"[{game_name}] Fetched: {data['name']} - ${total_price_str}")
        else:
            print(f"[{game_name}] Skipped: {card}")
        time.sleep(0.1)
    return results

def main():
    all_data = []
    
    # Process MTG
    all_data.extend(process_file('mtg_cards.txt', get_mtg_data, 'MTG'))
    
    # Process YGO
    all_data.extend(process_file('ygo_cards.txt', get_yugioh_data, 'YGO'))
    
    # Calculate Totals
    total_collection_value = 0.0
    total_profit_loss = 0.0
    
    for item in all_data:
        if item['sort_val'] > 0:
            total_collection_value += item['sort_val']
        if item['profit_loss'] is not None:
            total_profit_loss += item['profit_loss']
            
    # Sort by price
    all_data.sort(key=lambda x: x['sort_val'], reverse=True)

    # 1. Update Excel
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Card Prices"
        ws.append(['Date', 'Game', 'Card Name', 'Set', 'Prices', 'Notes'])
        
        today = datetime.now().strftime('%Y-%m-%d')
        for item in all_data:
            d = item['data']
            note = f"x{item['quantity']}" if item['quantity'] > 1 else ""
            if item['profit_loss'] is not None:
                sign = "+" if item['profit_loss'] >= 0 else ""
                note += f" (P/L: {sign}${item['profit_loss']:.2f})"
                
            ws.append([today, d['game'], d['name'], d['set'], item['price_str'], note])
            
        ws.append(['Total', '', '', '', f"{total_collection_value:.2f}", f"Total P/L: ${total_profit_loss:.2f}"])
        
        # Formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            
        wb.save('MY_COLLECTION_PRICES.xlsx')
        print("Saved to MY_COLLECTION_PRICES.xlsx")
    except Exception as e:
        print(f"Error saving Excel: {e}")

    # 2. Update History & Graph
    update_history_and_graph(total_collection_value)
    
    # 3. Generate HTML Report
    generate_html_report(all_data, total_collection_value, total_profit_loss)

if __name__ == "__main__":
    main()
